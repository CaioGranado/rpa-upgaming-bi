import logging
from pathlib import Path

import openpyxl
import pandas as pd

from config.settings import TXT_ORGANICOS

logger = logging.getLogger(__name__)


# ==============================================================================
# 1. FUNÇÕES DE LIMPEZA DE ARQUIVOS CRUS (Pós-Download)
# ==============================================================================
def organizar_datas(df, colunas_data, coluna_ordem=None):
    for col in colunas_data:
        if col in df.columns:
            limpo = df[col].astype(str).str.replace('.', '/').str.strip()
            limpo = limpo.replace(['nan', 'NaT', 'None', ''], pd.NA)
            # Removemos a conversão para texto (String) e mantemos como Datetime nativo
            # para o Excel reconhecer meses/dias corretamente e a dinâmica funcionar.
            df[col] = pd.to_datetime(limpo, dayfirst=True, errors='coerce')
    # Como por padrão nossa coluna_ordem=None/False, então o primeiro passo é verificar se tem algum valor, coluna_ordem=True
    # Depois se a mesma de fato estão dentro do nosso df
    if coluna_ordem and coluna_ordem in df.columns:
        df = df.sort_values(by=coluna_ordem, ascending=True)
        
    return df

def converter_para_numero(df, colunas):
    for col in colunas:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce')
    return df

def tratar_booleanos(df):
    colunas_bool = ['KYCStatus', 'MailVerified', 'RegistrationApproved', 'IsCellxpert', 'Locked', 'RequiresManualConfirmation']
    for col in colunas_bool:
        if col in df.columns:
            df[col] = df[col].astype(str).str.strip().str.lower().map({
                'true': 'True',
                'false': 'False',
                '1': 'True',
                '0': 'False',
                '1.0': 'True',
                '0.0': 'False'
            }).fillna('')
    return df
    
def tratar_ugs(df):
    df = converter_para_numero(df, ['Userid', 'Betcount', 'Wincount', 'BetAmount', 'WinAmount', 'Profit'])
    if 'ParentAffiliateUserName' in df.columns:
        df['ParentAffiliateUserName'] = df['ParentAffiliateUserName'].fillna(TXT_ORGANICOS)
    return df

def tratar_ftd(df):
    df = converter_para_numero(df, ['UserProfileID', 'LocalAmount', 'BrandAmount'])
    if 'ParentAffiliateUserName' in df.columns:
        df['ParentAffiliateUserName'] = df['ParentAffiliateUserName'].fillna(TXT_ORGANICOS)
    df = organizar_datas(df, ['RegistrationDate', 'TransactionDate'], 'TransactionDate')
    return df

def tratar_nc(df):
    df = converter_para_numero(df, ['UserProfileID', 'Mobile', 'TotalDepositBrandCurrency'])
    if 'ParentUserName' in df.columns:
        df['ParentUserName'] = df['ParentUserName'].fillna(TXT_ORGANICOS)
    df = organizar_datas(df, ['RegistrationDate', 'LastLoginedDate', 'BirthDate'], 'RegistrationDate')
    return df

def tratar_transacoes(df):
    df = converter_para_numero(df, ['TransactionId', 'UserProfileId', 'BrandAmountDecimal', 'LocalAmountDecimal'])
    df = organizar_datas(df, ['StartDate', 'EndDate'], 'StartDate')
    return df

def aplicar_regras_de_negocio(df: pd.DataFrame, nome_arquivo: str) -> pd.DataFrame:
    """Aplicando os tratamentos base para cada tipo de relatório"""
    df = tratar_booleanos(df)

    if "UGS" in nome_arquivo:
        return tratar_ugs(df)
    if "FTD" in nome_arquivo:
        return tratar_ftd(df)
    if "NC -" in nome_arquivo:
        return tratar_nc(df)
    if "Transações" in nome_arquivo:
        return tratar_transacoes(df)

    return df

def _processar_arquivo_unico(caminho: Path, colunas_moeda: list) -> bool:
    nome_arquivo = caminho.name

    colunas_data_geral = [
        'RegistrationDate', 'TransactionDate', 'LastLoginedDate',
        'BirthDate', 'StartDate', 'EndDate'
    ]

    try:
        df = pd.read_excel(caminho, engine='calamine')
        df = aplicar_regras_de_negocio(df, nome_arquivo)

        with pd.ExcelWriter(caminho, engine='openpyxl', datetime_format='dd/mm/yyyy hh:mm') as writer:            df.to_excel(writer, index=False)

        wb = openpyxl.load_workbook(caminho)
        try:
            ws = wb.active
            for col_idx, col_name in enumerate(df.columns, start=1):
                if col_name in colunas_moeda:
                    for row in range(2, ws.max_row + 1):
                        ws.cell(row=row, column=col_idx).number_format = '"R$ "#,##0.00'

                elif col_name in colunas_data_geral:
                    for row in range(2, ws.max_row + 1):
                        ws.cell(row=row, column=col_idx).number_format = r'dd\/mm\/yyyy hh:mm'

            wb.save(caminho)
        finally:
            wb.close()

        return True
    except Exception:
        logger.exception(f"Erro Crítico ao tratar o arquivo: {nome_arquivo}.")
        return False

def tratar_relatorios_crus(lista_arquivos):
    logger.info("Iniciando módulo de Transformação de Arquivos Crus (Pandas)...")
    arquivos_tratados = []
    colunas_moeda_geral = [
        'BetAmount', 'WinAmount', 'Profit', 'LocalAmount', 'BrandAmount', 
        'TotalDepositBrandCurrency', 'BrandAmountDecimal', 'LocalAmountDecimal'
    ]

    for arquivo in lista_arquivos:
        caminho = Path(arquivo)
        
        # Se for JSON, guarda na lista e pula o tratamento Pandas
        if caminho.suffix.lower() == ".json":
            logger.info(f" -> Pulando formatação: {caminho.name} é um JSON.")
            arquivos_tratados.append(str(caminho))
            continue
            
        # Se não for Excel (.xlsx ou .xls), ignora completamente (trava do desktop.ini)
        if caminho.suffix.lower() not in [".xlsx", ".xls"]:
            logger.debug(f" -> Ignorando arquivo de sistema/desconhecido: {caminho.name}")
            continue

        logger.info(f" -> Processando regras no arquivo: {caminho.name}")
        if _processar_arquivo_unico(caminho, colunas_moeda_geral):
            arquivos_tratados.append(str(caminho))
            
    logger.info("Transformação de arquivos crus concluída com sucesso.")
    return arquivos_tratados

# ==============================================================================
# 2. FUNÇÕES DE BLINDAGEM E PREPARAÇÃO PARA INJEÇÃO (Pré-Win32)
# ==============================================================================
def _blindar_pin(val):
    if pd.isna(val) or str(val).strip() == "": 
        return ""
    val_str = str(val).strip()
    return f"'{val_str.removesuffix('.0')}"

def _blindar_booleano(val):
    if pd.isna(val) or str(val).strip() == "":
        return ""
    return f"'{val}"

def blindar_dados(df: pd.DataFrame) -> pd.DataFrame:
    if 'Pin' in df.columns:
        df['Pin'] = df['Pin'].apply(_blindar_pin)

    colunas_bool = ['KYCStatus', 'MailVerified', 'RegistrationApproved', 'IsCellxpert', 'Locked', 'RequiresManualConfirmation']
    for col in colunas_bool:
        if col in df.columns:
            df[col] = df[col].apply(_blindar_booleano)
            
    # --- ESCUDO DEFINITIVO DE DATAS CONTRA BUGS DO EXCEL E WIN32COM ---
    colunas_data_geral = [
        'RegistrationDate', 'TransactionDate', 'LastLoginedDate', 
        'BirthDate', 'StartDate', 'EndDate'
    ]
    for col in df.columns:
        if col in colunas_data_geral:
            datas_temp = pd.to_datetime(df[col], dayfirst=True, errors='coerce')
            
            # 2. Converte para o texto ISO (AAAA-MM-DD). 
            # O Excel recebe isso e formata perfeitamente como data BR nativa, sem inverter meses!
            df[col] = datas_temp.dt.strftime('%Y-%m-%d %H:%M:%S')
            
            # 3. Limpa os vazios para evitar erros de injeção
            df[col] = df[col].fillna("")
            
    return df.fillna("")

def aplicar_corte_datas_futuras(df: pd.DataFrame, coluna_data: str, data_alvo) -> pd.DataFrame:
    if coluna_data in df.columns:
        limite_corte = pd.Timestamp(data_alvo.date()) + pd.Timedelta(days=1)

        datas_temporarias = pd.to_datetime(df[coluna_data], dayfirst=True, errors='coerce')
        eh_vazio_real = df[coluna_data].astype(str).str.strip().isin(['', 'nan', 'NaT', 'None', '<NA>'])

        df = df[(datas_temporarias < limite_corte) | eh_vazio_real]
        df = df.reset_index(drop=True)
    return df

def extrair_bloco_dinamica(df_source: pd.DataFrame, col_idx_dia: int, col_idx_valores: list, col_names: list) -> pd.DataFrame:
    mask = df_source.iloc[:, col_idx_dia].astype(str).str.contains('Rótulos|Dia', case=False)
    if not mask.any(): return pd.DataFrame()
        
    idx = df_source[mask].index[0]
    df = df_source.iloc[idx+1:].copy()
    
    df = df.iloc[:, [col_idx_dia] + col_idx_valores]
    df.columns = ['Dia_raw'] + col_names
    
    df['Dia'] = pd.to_numeric(df['Dia_raw'], errors='coerce')
    df = df.dropna(subset=['Dia'])
    df['Dia'] = df['Dia'].astype(int)
    df = df[df['Dia'] > 0]
    
    for col in col_names:
        df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
        
    return df[['Dia'] + col_names].copy()

def garantir_continuidade_temporal(df_extraido: pd.DataFrame, limite_dia: int) -> pd.DataFrame:
    """Vamos garantir a coerencia temporal, e preencher NaN com zero"""
    # 1. Auditoria e Alertas
    dias_extraidos = df_extraido['Dia'].tolist() if not df_extraido.empty else []
    dias_faltantes = [d for d in range(1, limite_dia + 1) if d not in dias_extraidos]

    if dias_faltantes:
        logging.getLogger(__name__).warning(f"ALERTA DE DADO AUSENTE: Há valores NaN, os identificados foram preenchidos com ZERO: {dias_faltantes}")
    else:
        logging.getLogger(__name__).info(f"Todos nossos dados estão preenchidos corretamente, indo do Dia 01 a {limite_dia:02d}")

    # 2. Merge
    dias_perfeitos = pd.DataFrame({'Dia': range(1, limite_dia + 1)})
    try:
        df_completo = pd.merge(dias_perfeitos, df_extraido, on='Dia', how='left', validate='1:1').fillna(0)
    except pd.errors.MergeError:
        logging.getLogger(__name__).exception("ERRO DE INTEGRIDADE: Dias duplicados encontrados na tabela! Detalhe do Pandas")
        raise
    #3. Tipagem Definitiva
    for col in df_completo.columns:
        if col == 'Dia':
            df_completo[col] = df_completo[col].astype(int)
        else:
            df_completo[col] = pd.to_numeric(df_completo[col], errors='coerce').fillna(0)   

    return df_completo 